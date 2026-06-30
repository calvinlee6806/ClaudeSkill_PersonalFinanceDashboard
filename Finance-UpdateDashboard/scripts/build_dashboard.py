"""
Build a self-contained HTML finance dashboard from FinanceDB.xlsx.

Usage:
    python build_dashboard.py <FinanceDB.xlsx> [output.html]

Reads the 'Database' and 'Trips' sheets, embeds the data into
../assets/dashboard_template.html, and writes a single offline-openable HTML file.
Requires: openpyxl  (pip install openpyxl)
"""
import sys
import json
import datetime
import pathlib

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
TEMPLATE = HERE.parent / "assets" / "dashboard_template.html"


def js_date(d):
    return d.strftime("%Y-%m-%d") if isinstance(d, datetime.datetime) else d


def num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    rows = list(wb["Database"].iter_rows(values_only=True))
    tx = []
    for r in rows[1:]:
        if not r[0]:  # skip blank rows (no TransactionID)
            continue
        tx.append({
            "date": js_date(r[1]), "orig": num(r[2]), "cur": r[3],
            "base": num(r[4]), "account": r[5], "merchant": r[7],
            "cat": r[8], "sub": r[9], "owner": r[10],
            "travel": bool(r[11]), "notable": bool(r[12]), "tags": r[13],
        })

    trips = []
    if "Trips" in wb.sheetnames:
        # row 0 = title, row 1 = header, data starts at row 2
        for r in list(wb["Trips"].iter_rows(values_only=True))[2:]:
            if r[0]:
                trips.append({"name": r[0], "start": js_date(r[1]),
                              "end": js_date(r[2]), "loc": r[3]})

    return {
        "tx": tx,
        "trips": trips,
        "generated": datetime.date.today().isoformat(),
        "base": "GBP",
    }


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "Finance_Dashboard.html"

    data = extract(xlsx)
    template = TEMPLATE.read_text(encoding="utf-8")
    html = template.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    pathlib.Path(out).write_text(html, encoding="utf-8")

    print(f"Wrote {out}  ({len(data['tx'])} transactions, {len(data['trips'])} trips)")


if __name__ == "__main__":
    main()
